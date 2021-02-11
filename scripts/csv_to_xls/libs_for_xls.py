
import pandas as pd
import xlsxwriter
import csv
import openpyxl
import xlrd, xlwt


# Write the DataFrame to csv
#df.to_csv("example.csv")


##### pandas #######

## конвертация из csv в xlsx или xls как есть (не добавляется нумерация строк
## и первая строка не форматируется в заголовок таблицы)
## Делает xls или xlsx файлы по расширению
## можно создавать несколько листов
##сделаем датафрейм из csv, учитывая, что шапка в исходном файле не предусмотрена
# df = pd.read_csv("example.csv", header=None)
# print(df)
# of = pd.ExcelWriter('exampleW.xls') #cоздадим объект выходного файла xls
# df.to_excel(of, 'Sheet1', index=False, header=False) # запишем имеющийся датафрейм в файл без шапки и индекса
# df.to_excel(of, 'Sheet2', index=True, header=True)
# of.save()  # и сохраним его


# ##чтение из xls xlsx
# xl = pd.ExcelFile('example.xls')   #загружаем файл
# print(xl.sheet_names)               #печатаем список названий листов
# df1 = xl.parse(xl.sheet_names[0])   #загружает в датафрейм содержимое первого листа
# print(df1)


## конвертация из csv в xlsx - добавляется нумерация строк
## и первая строка форматируется в заголовок таблицы
## Не делает xls файлы даже при указании расширения (все равно файл внутри xlsx)
## можно делать документ на несколько листов
# df = pd.read_csv("example.csv")       #сделает датафрейм из csv
# print(df)
# of = pd.ExcelWriter('exampleW.xlsx', engine='xlsxwriter') #cоздадим объект выходного файла xlsx
# df.to_excel(of, 'Sheet1')       # запишем имеющийся датафрейм в файл и сохраним его
# of.save()

## конвертация из csv в xlsx или xls- добавляется нумерация строк
## и первая строка форматируется в заголовок таблицы
## Делает xls файлы при указании расширения
# df = pd.read_csv("example.csv")       #сделает датафрейм из csv
# print(df)
# of = pd.ExcelWriter('exampleW.xls') #cоздадим объект выходного файла xls
# df.to_excel(of)       # запишем имеющийся датафрейм в файл и сохраним его
# of.save()


## конвертация из csv в xlsx - добавляется нумерация строк
## и первая строка форматируется в заголовок таблицы
## Не делает xls файлы даже при указании расширения (все равно файл внутри xlsx)
# df = pd.read_csv("example.csv")       #сделает датафрейм из csv
# print(df)
# #Чтобы записать Pandas DataFrames обратно в файл Excel,
# # можно использовать функцию dataframe_to_rows () из модуля utils:
# from openpyxl.utils.dataframe import dataframe_to_rows
# # Initialize a workbook
# wb = openpyxl.Workbook()
# # Get the worksheet in the active workbook
# ws = wb.active
# #выведем датафрейм построчно
# for r in dataframe_to_rows(df, index=False, header=True):
#     print(r)
# # запишем строки датафрейма в наш рабочий лист
# for r in dataframe_to_rows(df, index=False, header=True):
#     ws.append(r)
# #print(ws['C5'].value)
# wb.save('new.xlsx')



# ##### openpyxl #######
#
# # не поддерживает старый xls. Пишет: openpyxl does not support the old .xls file format,
# # please use xlrd to read this file, or convert it to the more recent .xlsx file format.
#
# # Функция load_workbook() принимает имя файла в качестве аргумента и возвращает объект рабочей книги
# wb = openpyxl.load_workbook('example.xlsx')
# print(type(wb))
# # получим список названий листов
# print(wb.get_sheet_names())
#
# # получим лист по имени
# sheet = wb.get_sheet_by_name('Лист2')
# print(type(sheet))
# # выведем имя листа
# print(sheet.title)
# # а какой лист сейчас активен в документе?
# print(wb.active)
#
# # Retrieve the value of a certain cell
# print(sheet['A1'].value)
# # Select element 'B2' of your sheet
# c = sheet['B2']
# # Retrieve the row number of your element
# print(c.row)
# # Retrieve the column letter of your element
# print(c.column)
# # Retrieve the coordinates of the cell
# print(c.coordinate)
# # Вы также можете получить значения ячеек с помощью функции cell ().
# # Передайте аргументы row и column, добавьте значения к этим аргументам,
# # которые соответствуют значениям ячейки, которые вы хотите получить,
# # и, конечно же, не забудьте добавить атрибут value:
# print(sheet.cell(row=1, column=2).value)
#
# # след-е строки возращают букву колонки по номеру и наоборот
# print(openpyxl.utils.get_column_letter(1))
# print(openpyxl.utils.column_index_from_string('A'))
#
# # построчный вывод значений из области
# for cellObj in sheet['A1':'C3']:
#       for cell in cellObj:
#               print(cell.coordinate, cell.value)
#       print('--- END ---')
#
# # сколько в нашем листе колонок и строк?
# print(sheet.max_row)
# print(sheet.max_column)


# #??????????
# #Openpyxl имеет поддержку Pandas DataFrames.
# # И можно использовать функцию DataFrame ()
# # из пакета Pandas, чтобы поместить значения листа в DataFrame:
# # Convert Sheet to DataFrame
# df = pd.DataFrame(sheet.values)
# #Если вы хотите указать заголовки и индексы, вам нужно добавить немного больше кода:
# # Put the sheet values in `data`
# data = sheet.values
# # Indicate the columns in the sheet values
# cols = next(data)[1:]
# # Convert your data to a list
# data = list(data)
# # Read in the data at index 0 for the indices
# idx = [r[0] for r in data]
# # Slice the data at index 1
# data = (islice(r, 1, None) for r in data)
# # Make your DataFrame
# df = pd.DataFrame(data, index=idx, columns=cols)
# #??????????



########xlrd/xlwt############## (+ xlutils - основана на предыдущих двух)

# Чтение
# workbook = xlrd.open_workbook('example.xls')
# # # Loads only current sheets to memory
# # workbook = xlrd.open_workbook('example.xls', on_demand = True)
# # Если вы не хотите рассматривать всю книгу,
# # можно использовать такие функции, как
# # sheet_by_name () или sheet_by_index (),
# # чтобы извлекать листы, которые необходимо использовать в анализе.
# worksheet = workbook.sheet_by_name('Лист1')
# worksheet1 = workbook.sheet_by_index(0)
# # вернуть значение ячейки с координатами
# print(worksheet.cell(1, 1).value)

# Запись
# # Поячеечно
# book = xlwt.Workbook(encoding="utf-8")
# # Add a sheet to the workbook
# sheet1 = book.add_sheet("Python Sheet 1")
# # Write to the sheet of the workbook
# sheet1.write(0, 0, "This is the First Cell of the First Sheet")
# # Save the workbook
# book.save("spreadsheet.xls")


# # ????????????
# # какой-то сложный способ писать поячеечно
# # Initialize a workbook
# book = xlwt.Workbook()
# # Add a sheet to the workbook
# sheet1 = book.add_sheet("Sheet1")
# # The data
# cols = ["A", "B", "C", "D", "E"]
# txt = [0,1,2,3,4]
# # Loop over the rows and columns and fill in the values
# for num in range(5):
#       row = sheet1.row(num)
#       for index, col in enumerate(cols):
#           value = txt[index] + num
#           row.write(index, value)
# # Save the result
# book.save("test.xls")
# # ????????????


######## pyexcel ###########

import pyexcel
from pyexcel._compact import OrderedDict
# Чтение
# Чтобы получить данные в массиве, можно использовать функцию
# get_array (), которая содержится в пакете pyexcel:
my_array = pyexcel.get_array(file_name="example.xls")
# print(my_array)

# Также можно получить данные в упорядоченном словаре списков,
# используя функцию get_dict ():
my_dict = pyexcel.get_dict(file_name="example.xls", name_columns_by_row=0)
# print(my_dict)

# Тоже упорядоченный словарь, только элементы словаря - листы
book_dict = pyexcel.get_book_dict(file_name="example.xls")
# print(book_dict)

# Наконец, вы можете просто получить записи с pyexcel благодаря функции
# get_records (). Просто передайте аргумент file_name функции и обратно получите список словарей:
records = pyexcel.get_records(file_name="example.xls")
# print(records)

##Запись
# # можно также легко экспортировать массивы обратно в электронную таблицу.
# # Для этого используется функция save_as ()
# # с передачей массива и имени целевого файла в аргумент dest_file_name:
# data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
# pyexcel.save_as(array=data, dest_file_name="array_data.xls")

# # Однако, если у вас есть словарь, нужно будет использовать функцию
# # save_book_as (). Передайте двумерный словарь в bookdict и укажите имя файла
# array_dictionary = OrderedDict([('Sheet 1', [
#                                    ['ID', 'AGE', 'SCORE'],
#                                    [1, 22, 5],
#                                    [2, 15, 6],
#                                    [3, 28, 9]
#                                   ]),
#                                 ('Sheet 2', [
#                                     ['X', 'Y', 'Z'],
#                                     [1, 2, 3],
#                                     [4, 5, 6],
#                                     [7, 8, 9]
#                                   ]),
#                                 ('Sheet 3', [
#                                     ['M', 'N', 'O', 'P'],
#                                     [10, 11, 12, 13],
#                                     [14, 15, 16, 17],
#                                     [18, 19, 20, 21]
#                                    ])])
# pyexcel.save_book_as(bookdict=array_dictionary, dest_file_name="2d_array_data.xls")

######## csv ############
# # Чтение
# for row in csv.reader(open('example.csv'), delimiter=','):
#     print(row)
#
# # Запись
# data = [[1, 2, 3], [4, 5, 6], [7, 8, 9]]
# outfile = open('data.csv', 'w')
# writer = csv.writer(outfile, delimiter=';', quotechar='"')
# writer.writerows(data)
# outfile.close()



if __name__ == '__main__':
    import sys
    print('Using version:', sys.version[:5])